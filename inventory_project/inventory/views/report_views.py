from django.http import HttpResponse


def reports(request, section):
    """Временная заглушка для отчетов"""
    return HttpResponse(f"Страница отчетов: {section} (в разработке)")


from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.db.models import Sum
from ..models import Doc, Detail, Nom, Postav, Obct, Podraz, Fio, Category
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, timedelta


def incom_report(request):
    """Отчет о поступлении ТМЦ"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'doc')  # 'doc' или 'created'

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_incom_report_table(request, date_start, date_end, date_type)

    context = {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    }
    return render(request, 'inventory/reports/incom_report.html', context)


def render_incom_report_table(request, date_start, date_end, date_type='doc'):
    """Рендеринг таблицы отчета"""

    # Базовый запрос
    docs = Doc.objects.filter(oper=2).select_related('postav').prefetch_related('details', 'details__id_nom',
                                                                                'details__id_nom__izm')

    # Фильтрация по дате в зависимости от выбранного типа
    if date_start:
        if date_type == 'doc':
            docs = docs.filter(datadoc__gte=date_start)
        else:  # 'created'
            # Преобразуем дату в начало дня
            start_datetime = datetime.strptime(date_start, '%Y-%m-%d')
            docs = docs.filter(update_date__gte=start_datetime)

    if date_end:
        if date_type == 'doc':
            docs = docs.filter(datadoc__lte=date_end)
        else:  # 'created'
            # Преобразуем дату в конец дня
            end_datetime = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
            docs = docs.filter(update_date__lt=end_datetime)

    docs = docs.order_by('-datadoc')

    # Добавляем агрегированные суммы для каждого документа
    docs_with_totals = []
    grand_total = 0
    grand_total_vat = 0
    grand_total_without_vat = 0

    for doc in docs:
        details = list(doc.details.all())
        total_without_vat = sum(d.cost for d in details)
        total_vat = sum(d.vat_amount for d in details)
        total = sum(d.total_with_vat for d in details)

        grand_total_without_vat += total_without_vat
        grand_total_vat += total_vat
        grand_total += total

        # Форматируем даты для отображения
        date_start_display = ''
        date_end_display = ''

        if date_start:
            try:
                date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
            except:
                date_start_display = date_start

        if date_end:
            try:
                date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
            except:
                date_end_display = date_end

        docs_with_totals.append({
            'id': doc.id,
            'nomer': doc.nomer,
            'datadoc': doc.datadoc,
            'postav': doc.postav,
            'details': details,
            'total': total,
            'total_without_vat': total_without_vat,
            'total_vat': total_vat,
            'created_at': doc.update_date,  # добавляем дату создания
        })

    html = render_to_string('inventory/reports/incom_report_table.html', {
        'docs': docs_with_totals,
        'date_start': date_start_display if 'date_start_display' in locals() else '',
        'date_end': date_end_display if 'date_end_display' in locals() else '',
        'date_type': date_type,
        'grand_total': grand_total,
        'grand_total_vat': grand_total_vat,
        'grand_total_without_vat': grand_total_without_vat
    })
    return HttpResponse(html)


def incom_report_excel(request):
    """Экспорт отчета в Excel"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'doc')

    docs = Doc.objects.filter(oper=2).select_related('postav').prefetch_related('details', 'details__id_nom',
                                                                                'details__id_nom__izm')

    if date_start:
        if date_type == 'doc':
            docs = docs.filter(datadoc__gte=date_start)
        else:
            start_datetime = datetime.strptime(date_start, '%Y-%m-%d')
            docs = docs.filter(update_date__gte=start_datetime)

    if date_end:
        if date_type == 'doc':
            docs = docs.filter(datadoc__lte=date_end)
        else:
            end_datetime = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
            docs = docs.filter(update_date__lt=end_datetime)

    docs = docs.order_by('-datadoc')

    # Создаем Excel файл
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по поступлению"

    # Заголовки
    headers = ['№ п/п', 'Документ', 'Дата', 'Поставщик', 'Наименование', 'Ед.изм.', 'Кол-во', 'Цена',
               'Стоимость без НДС', 'НДС %', 'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for doc in docs:
        for detail in doc.details.all():
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=doc.nomer)
            ws.cell(row=row_num, column=3, value=doc.datadoc.strftime('%d.%m.%Y'))
            ws.cell(row=row_num, column=4, value=doc.postav.title if doc.postav else '')
            ws.cell(row=row_num, column=5, value=detail.id_nom.title if detail.id_nom else '')
            ws.cell(row=row_num, column=6, value=detail.id_nom.izm.title if detail.id_nom and detail.id_nom.izm else '')
            ws.cell(row=row_num, column=7, value=float(detail.kolvo))
            ws.cell(row=row_num, column=8, value=float(detail.price))
            ws.cell(row=row_num, column=9, value=float(detail.cost))
            ws.cell(row=row_num, column=10, value=float(detail.vat_rate))
            ws.cell(row=row_num, column=11, value=float(detail.vat_amount))
            ws.cell(row=row_num, column=12, value=float(detail.total_with_vat))
            row_num += 1
        row_num += 1  # пустая строка после документа

    # Настраиваем ширину колонок
    for col in range(1, 13):
        ws.column_dimensions[chr(64 + col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=incom_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def reports_menu(request, section):
    """Страница меню отчетов"""
    # Просто показываем меню, независимо от section
    reports = {
        'incom': 'Поступление ТМЦ',
        'move': 'Перемещение ТМЦ',
        'remain': 'Наличие ТМЦ',
        'suppliers':'По поставщикам',
        'objects': 'По объектам',
        'departments':'По подразделениям',
        'fio': 'По подотчетным лицам',
        'categories':'По категориям',
        'top10': 'ТОП-10'


    }

    context = {
        'data': reports,
        'title': 'Отчеты',
        'section': section,
        'kind': 'reports',
    }
    return render(request, 'inventory/menu.html', context)


def move_report(request):
    """Отчет о перемещении ТМЦ"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_move_report_table(request, date_start, date_end, date_type)

    context = {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    }
    return render(request, 'inventory/reports/move_report.html', context)


def render_move_report_table(request, date_start, date_end, date_type='created'):
    """Рендеринг таблицы отчета по перемещению"""
    from datetime import datetime, timedelta

    # Объявляем переменные ДО условий
    date_start_display = ''
    date_end_display = ''

    docs = Doc.objects.filter(oper=3).select_related('fio', 'obct', 'obct__idpodraz').prefetch_related('details',
                                                                                                       'details__id_nom',
                                                                                                       'details__id_nom__izm')

    if date_start:
        if date_type == 'doc':
            docs = docs.filter(datadoc__gte=date_start)
        else:
            start_datetime = datetime.strptime(date_start, '%Y-%m-%d')
            docs = docs.filter(update_date__gte=start_datetime)

        # Форматируем дату для отображения
        try:
            date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_start_display = date_start

    if date_end:
        if date_type == 'doc':
            docs = docs.filter(datadoc__lte=date_end)
        else:
            end_datetime = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
            docs = docs.filter(update_date__lt=end_datetime)

        # Форматируем дату для отображения
        try:
            date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_end_display = date_end

    docs = docs.order_by('-datadoc')

    docs_with_totals = []
    grand_total = 0
    grand_total_vat = 0
    grand_total_without_vat = 0

    for doc in docs:
        details = list(doc.details.all())
        total_without_vat = sum(d.cost for d in details)
        total_vat = sum(d.vat_amount for d in details)
        total = sum(d.total_with_vat for d in details)

        grand_total_without_vat += total_without_vat
        grand_total_vat += total_vat
        grand_total += total

        docs_with_totals.append({
            'id': doc.id,
            'nomer': doc.nomer,
            'datadoc': doc.datadoc,
            'fio': doc.fio.title if doc.fio else '—',
            'obct': doc.obct.title if doc.obct else '—',
            'podraz': doc.obct.idpodraz.title if doc.obct and doc.obct.idpodraz else '—',
            'details': details,
            'total': total,
            'total_without_vat': total_without_vat,
            'total_vat': total_vat,
        })

    html = render_to_string('inventory/reports/move_report_table.html', {
        'docs': docs_with_totals,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'date_type': date_type,
        'grand_total': grand_total,
        'grand_total_vat': grand_total_vat,
        'grand_total_without_vat': grand_total_without_vat,
    })
    return HttpResponse(html)

def move_report_excel(request):
    """Экспорт отчета по перемещению в Excel"""
    from datetime import datetime, timedelta
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    docs = Doc.objects.filter(oper=3).select_related('fio', 'obct', 'obct__idpodraz').prefetch_related('details', 'details__id_nom', 'details__id_nom__izm')

    if date_start:
        if date_type == 'doc':
            docs = docs.filter(datadoc__gte=date_start)
        else:
            start_datetime = datetime.strptime(date_start, '%Y-%m-%d')
            docs = docs.filter(update_date__gte=start_datetime)

    if date_end:
        if date_type == 'doc':
            docs = docs.filter(datadoc__lte=date_end)
        else:
            end_datetime = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
            docs = docs.filter(update_date__lt=end_datetime)

    docs = docs.order_by('-datadoc')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по перемещению"

    headers = ['№ п/п', 'Документ', 'Дата', 'Подразделение', 'Объект', 'Подотчет',
               'Наименование', 'Ед.изм.', 'Кол-во', 'Цена', 'Стоимость без НДС', 'НДС %', 'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for doc in docs:
        for detail in doc.details.all():
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=doc.nomer)
            ws.cell(row=row_num, column=3, value=doc.datadoc.strftime('%d.%m.%Y'))
            ws.cell(row=row_num, column=4, value=doc.obct.idpodraz.title if doc.obct and doc.obct.idpodraz else '')
            ws.cell(row=row_num, column=5, value=doc.obct.title if doc.obct else '')
            ws.cell(row=row_num, column=6, value=doc.fio.title if doc.fio else '')
            ws.cell(row=row_num, column=7, value=detail.id_nom.title if detail.id_nom else '')
            ws.cell(row=row_num, column=8, value=detail.id_nom.izm.title if detail.id_nom and detail.id_nom.izm else '')
            ws.cell(row=row_num, column=9, value=abs(float(detail.kolvo)))
            ws.cell(row=row_num, column=10, value=abs(float(detail.price)))
            ws.cell(row=row_num, column=11, value=abs(float(detail.cost)))
            ws.cell(row=row_num, column=12, value=abs(float(detail.vat_rate)))
            ws.cell(row=row_num, column=13, value=abs(float(detail.vat_amount)))
            ws.cell(row=row_num, column=14, value=abs(float(detail.total_with_vat)))
            row_num += 1
        row_num += 1

    for col in range(1, 15):
        ws.column_dimensions[chr(64 + col)].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=move_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def remain_report(request):
    """Отчет об остатках ТМЦ на текущую дату"""
    show_zero = request.GET.get('show_zero', 'all')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_remain_report_table(request, show_zero)

    return render(request, 'inventory/reports/remain_report.html')


def render_remain_report_table(request, show_zero='all'):
    """Рендеринг таблицы отчета об остатках (группировка по товару+цена с объединением)"""
    from django.db import connection

    # Убираем HAVING — получаем ВСЕ товары, включая нулевые остатки
    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            SUM(d.kolvo) as quantity
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper IN (1, 2, 4, 3, 5)
        GROUP BY n.id, n.title, i.title, d.price
        ORDER BY n.title, d.price
    """

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

        # Временный словарь для объединения одинаковых товаров с одинаковой ценой
        temp_dict = {}

        for row in rows:
            data = dict(zip(columns, row))
            quantity = float(data['quantity'])

            # Ключ: товар + цена
            key = f"{data['nom_id']}_{data['price']}"

            if key in temp_dict:
                # Если уже есть — суммируем количество
                temp_dict[key]['quantity'] += quantity
            else:
                # Если нет — добавляем новую запись
                temp_dict[key] = {
                    'nom_id': data['nom_id'],
                    'title': data['title'] or 'Без названия',
                    'izm': data['izm'] or 'шт',
                    'price': float(data['price']),
                    'quantity': quantity
                }

        # Формируем итоговый список с фильтрацией
        remains = []
        total_quantity = 0
        total_sum = 0

        for key, item in temp_dict.items():
            quantity = round(item['quantity'], 0)

            # Фильтрация по show_zero (уже на уровне Python)
            if show_zero == 'positive' and quantity <= 0:
                continue

            remains.append({
                'nom_id': item['nom_id'],
                'title': item['title'],
                'izm': item['izm'],
                'price': item['price'],
                'quantity': quantity
            })
            total_quantity += quantity if quantity > 0 else 0
            total_sum += quantity * item['price'] if quantity > 0 else 0

        # Сортируем по наименованию, затем по цене
        remains.sort(key=lambda x: (x['title'], x['price']))

        html = render_to_string('inventory/reports/remain_report_table.html', {
            'remains': remains,
            'total_quantity': total_quantity,
            'total_sum': total_sum,
        })
        return HttpResponse(html)

    except Exception as e:
        print(f"Ошибка в remain_report: {e}")
        import traceback
        traceback.print_exc()
        return HttpResponse('<div class="alert alert-danger">Ошибка формирования отчета</div>')

def remain_report_excel(request):
    """Экспорт отчета об остатках в Excel"""
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from datetime import datetime

    show_zero = request.GET.get('show_zero', 'all')

    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            SUM(d.kolvo) as quantity
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper IN (1, 2, 4, 3, 5)
        GROUP BY n.id, n.title, i.title, d.price
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

    # Объединяем одинаковые товары с одинаковой ценой
    temp_dict = {}
    for row in rows:
        nom_id = row[0]
        title = row[1] or 'Без названия'
        izm = row[2] or 'шт'
        price = float(row[3])
        quantity = float(row[4])

        key = f"{nom_id}_{price}"
        if key in temp_dict:
            temp_dict[key]['quantity'] += quantity
        else:
            temp_dict[key] = {
                'nom_id': nom_id,
                'title': title,
                'izm': izm,
                'price': price,
                'quantity': quantity
            }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Остатки ТМЦ"

    # Заголовки
    headers = ['№ п/п', 'Наименование', 'Ед.изм.', 'Цена', 'Количество']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    total_quantity = 0
    total_sum = 0
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    i = 1
    for key, item in temp_dict.items():
        quantity = round(item['quantity'], 0)

        if show_zero == 'positive' and quantity <= 0:
            continue

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=item['title'])
        ws.cell(row=row_num, column=3, value=item['izm'])
        ws.cell(row=row_num, column=4, value=item['price'])
        ws.cell(row=row_num, column=5, value=quantity)

        # Подсветка нулевых
        if quantity == 0:
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = yellow_fill

        total_quantity += quantity if quantity > 0 else 0
        total_sum += quantity * item['price'] if quantity > 0 else 0
        row_num += 1
        i += 1

    # Итоговая строка
    ws.cell(row=row_num, column=4, value='ИТОГО:')
    ws.cell(row=row_num, column=5, value=total_quantity)
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    ws.cell(row=row_num, column=5).font = Font(bold=True)

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=remain_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def nom_card(request, nom_id):
    """Карточка товара — история всех движений"""
    from decimal import Decimal

    nom = get_object_or_404(Nom, id=nom_id)

    # Получаем все движения по товару в хронологическом порядке
    movements = Detail.objects.filter(
        id_nom_id=nom_id
    ).select_related('id_doc', 'id_doc__postav', 'id_doc__fio', 'id_doc__obct').order_by('id_doc__datadoc', 'id')

    movement_list = []
    balance = Decimal('0')
    total_income = Decimal('0')
    total_outcome = Decimal('0')
    total_sum_without_vat = Decimal('0')
    total_vat = Decimal('0')
    total_sum = Decimal('0')

    for detail in movements:
        doc = detail.id_doc
        if not doc:
            continue

        kolvo = detail.kolvo

        # Обновляем остаток
        balance += kolvo

        # Определяем "от кого получено / кому отпущено"
        from_party = ''
        to_party = ''

        if doc.oper == 2:  # Приход
            from_party = doc.postav.title if doc.postav else '—'
            to_party = ''
        elif doc.oper == 3:  # Перемещение
            from_party = ''
            to_party = f"{doc.fio.title if doc.fio else '—'} на {doc.obct.title if doc.obct else '—'}"
        elif doc.oper == 4:  # Возврат
            from_party = doc.postav.title if doc.postav else '—'
            to_party = ''
        elif doc.oper == 5:  # Списание
            from_party = ''
            to_party = f"Списание: {doc.fio.title if doc.fio else '—'}"
        else:
            from_party = '—'
            to_party = '—'

        # Используем правильные поля для отображения
        cost_without_vat = abs(detail.cost) if detail.cost else Decimal('0')
        vat_amount = abs(detail.vat_amount) if detail.vat_amount else Decimal('0')
        total_with_vat = abs(detail.total_with_vat) if detail.total_with_vat else cost_without_vat + vat_amount

        movement_list.append({
            'doc_id': doc.id,
            'nomer': doc.nomer,
            'datadoc': doc.datadoc,
            'oper': doc.oper,
            'kolvo': kolvo,
            'from_party': from_party,
            'to_party': to_party,
            'price': abs(detail.price) if detail.price else Decimal('0'),
            'cost': cost_without_vat,
            'vat_rate': detail.vat_rate if detail.vat_rate else Decimal('0'),
            'vat_amount': vat_amount,
            'total_with_vat': total_with_vat,
            'balance_after': balance
        })

        if kolvo > 0:
            total_income += abs(kolvo)
        else:
            total_outcome += abs(kolvo)

        total_sum_without_vat += cost_without_vat
        total_vat += vat_amount
        total_sum += total_with_vat

    context = {
        'nom': nom,
        'movements': movement_list,
        'current_balance': balance,
        'total_income': total_income,
        'total_outcome': total_outcome,
        'total_sum_without_vat': total_sum_without_vat,
        'total_vat': total_vat,
        'total_sum': total_sum,
    }
    return render(request, 'inventory/reports/nom_card.html', context)


def suppliers_report(request):
    """Отчет по поставщикам"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_suppliers_report_table(request, date_start, date_end, date_type)

    return render(request, 'inventory/reports/supplier_report.html', {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    })


def render_suppliers_report_table(request, date_start, date_end, date_type='created'):
    """Рендеринг таблицы отчета по поставщикам"""
    from datetime import datetime, timedelta

    # Базовый запрос
    docs = Doc.objects.filter(oper=2).select_related('postav')

    # Фильтрация по дате
    if date_start:
        if date_type == 'doc':
            docs = docs.filter(datadoc__gte=date_start)
        else:
            start_datetime = datetime.strptime(date_start, '%Y-%m-%d')
            docs = docs.filter(update_date__gte=start_datetime)

    if date_end:
        if date_type == 'doc':
            docs = docs.filter(datadoc__lte=date_end)
        else:
            end_datetime = datetime.strptime(date_end, '%Y-%m-%d') + timedelta(days=1)
            docs = docs.filter(update_date__lt=end_datetime)

    # Группировка по поставщикам
    suppliers_dict = {}
    for doc in docs:
        if not doc.postav:
            continue
        postav_id = doc.postav.id
        if postav_id not in suppliers_dict:
            suppliers_dict[postav_id] = {
                'id': postav_id,
                'title': doc.postav.title,
                'doc_count': 0,
                'total': 0
            }
        suppliers_dict[postav_id]['doc_count'] += 1
        suppliers_dict[postav_id]['total'] += doc.total or 0

    suppliers = list(suppliers_dict.values())
    suppliers.sort(key=lambda x: x['title'])
    grand_total = sum(s['total'] for s in suppliers)

    # Форматирование дат
    date_start_display = ''
    date_end_display = ''
    if date_start:
        try:
            date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_start_display = date_start
    if date_end:
        try:
            date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_end_display = date_end

    html = render_to_string('inventory/reports/supplier_report_table.html', {
        'suppliers': suppliers,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'date_type': date_type,
        'grand_total': grand_total,
    })
    return HttpResponse(html)


def supplier_details(request, supplier_id):
    """Детализация по поставщику"""
    from datetime import datetime, timedelta

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    supplier = get_object_or_404(Postav, id=supplier_id)

    docs = Doc.objects.filter(oper=2, postav_id=supplier_id)

    # Фильтрация по дате (с правильным форматом)
    if date_start:
        try:
            # Парсим дату в формате DD.MM.YYYY
            start_date_obj = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date_obj = None

        if date_type == 'doc' and start_date_obj:
            docs = docs.filter(datadoc__gte=start_date_obj)
        elif start_date_obj:
            docs = docs.filter(update_date__gte=start_date_obj)

    if date_end:
        try:
            end_date_obj = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date_obj = None

        if date_type == 'doc' and end_date_obj:
            docs = docs.filter(datadoc__lte=end_date_obj)
        elif end_date_obj:
            end_datetime = datetime.combine(end_date_obj, datetime.max.time())
            docs = docs.filter(update_date__lte=end_datetime)

    docs = docs.order_by('-datadoc')

    # Собираем документы с деталями
    docs_with_details = []
    grand_total = 0

    for doc in docs:
        details = list(doc.details.all())
        total_without_vat = sum(d.cost for d in details)
        total_vat = sum(d.vat_amount for d in details)
        total = sum(d.total_with_vat for d in details)

        docs_with_details.append({
            'id': doc.id,
            'nomer': doc.nomer,
            'datadoc': doc.datadoc,
            'details': details,
            'total': total,
            'total_without_vat': total_without_vat,
            'total_vat': total_vat,
        })
        grand_total += total

    # Форматирование дат для отображения (они уже в правильном формате)
    context = {
        'supplier': supplier,
        'docs': docs_with_details,
        'date_start': date_start,
        'date_end': date_end,
        'grand_total': grand_total,
    }
    return render(request, 'inventory/reports/supplier_details.html', context)


def suppliers_report_excel(request):
    """Экспорт отчета по поставщикам в Excel"""
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    # Базовый запрос
    docs = Doc.objects.filter(oper=2).select_related('postav')

    # Фильтрация по дате
    if date_start:
        try:
            start_date_obj = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date_obj = None

        if date_type == 'doc' and start_date_obj:
            docs = docs.filter(datadoc__gte=start_date_obj)
        elif start_date_obj:
            docs = docs.filter(update_date__gte=start_date_obj)

    if date_end:
        try:
            end_date_obj = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date_obj = None

        if date_type == 'doc' and end_date_obj:
            docs = docs.filter(datadoc__lte=end_date_obj)
        elif end_date_obj:
            end_datetime = datetime.combine(end_date_obj, datetime.max.time())
            docs = docs.filter(update_date__lte=end_datetime)

    # Группировка по поставщикам
    suppliers_dict = {}
    for doc in docs:
        if not doc.postav:
            continue
        postav_id = doc.postav.id
        if postav_id not in suppliers_dict:
            suppliers_dict[postav_id] = {
                'title': doc.postav.title,
                'doc_count': 0,
                'total': 0
            }
        suppliers_dict[postav_id]['doc_count'] += 1
        suppliers_dict[postav_id]['total'] += doc.total or 0

    suppliers = list(suppliers_dict.values())
    suppliers.sort(key=lambda x: x['title'])
    grand_total = sum(s['total'] for s in suppliers)

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по поставщикам"

    # Заголовки
    headers = ['№ п/п', 'Поставщик', 'Количество документов', 'Общая сумма, руб.']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Заполняем данные
    row_num = 2
    for i, supplier in enumerate(suppliers, 1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=supplier['title'])
        ws.cell(row=row_num, column=3, value=supplier['doc_count'])
        ws.cell(row=row_num, column=4, value=supplier['total'])
        row_num += 1

    # Итоговая строка
    ws.cell(row=row_num, column=3, value='ИТОГО:')
    ws.cell(row=row_num, column=4, value=grand_total)
    ws.cell(row=row_num, column=3).font = Font(bold=True)
    ws.cell(row=row_num, column=4).font = Font(bold=True)

    # Настраиваем ширину колонок
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=supplier_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def objects_report(request):
    """Отчет по объектам (перемещения)"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_objects_report_table(request, date_start, date_end, date_type)

    return render(request, 'inventory/reports/object_report.html', {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    })


def render_objects_report_table(request, date_start, date_end, date_type='created'):
    """Рендеринг таблицы отчета по объектам"""
    from datetime import datetime, timedelta

    # Базовый запрос
    docs = Doc.objects.filter(oper=3).select_related('obct', 'obct__idpodraz')

    # Фильтрация по дате
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            docs = docs.filter(datadoc__gte=start_date)
        elif start_date:
            docs = docs.filter(update_date__date__gte=start_date)

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            docs = docs.filter(datadoc__lte=end_date)
        elif end_date:
            docs = docs.filter(update_date__date__lte=end_date)

    # Группировка по объектам
    objects_dict = {}
    for doc in docs:
        if not doc.obct:
            continue
        obct_id = doc.obct.id
        if obct_id not in objects_dict:
            objects_dict[obct_id] = {
                'id': obct_id,
                'title': doc.obct.title or 'Без названия',
                'podraz': doc.obct.idpodraz.title if doc.obct.idpodraz else '—',
                'total_without_vat': 0,
                'total_vat': 0,
                'total': 0
            }

        # Суммируем по документам
        for detail in doc.details.all():
            objects_dict[obct_id]['total_without_vat'] += abs(detail.cost)
            objects_dict[obct_id]['total_vat'] += abs(detail.vat_amount)
            objects_dict[obct_id]['total'] += abs(detail.total_with_vat)

    objects = list(objects_dict.values())
    objects.sort(key=lambda x: x['title'])

    grand_total = sum(o['total'] for o in objects)
    grand_total_without_vat = sum(o['total_without_vat'] for o in objects)
    grand_total_vat = sum(o['total_vat'] for o in objects)

    # Форматирование дат
    date_start_display = ''
    date_end_display = ''
    if date_start:
        try:
            date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_start_display = date_start
    if date_end:
        try:
            date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_end_display = date_end

    html = render_to_string('inventory/reports/object_report_table.html', {
        'objects': objects,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'date_type': date_type,
        'grand_total': grand_total,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
    })
    return HttpResponse(html)


def object_details(request, object_id):
    """Детализация по объекту — список товаров, переданных на объект"""
    from datetime import datetime, timedelta
    from django.db import connection

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    obj = get_object_or_404(Obct, id=object_id)

    # Базовый запрос — все строки перемещений на этот объект
    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum,
            MIN(d.id_doc) as doc_id
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 3 AND doc.obct_id = %s
    """

    params = [object_id]

    # Фильтрация по дате
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    items = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        data = dict(zip(columns, row))
        quantity = float(data['quantity'])
        total_cost = float(data['total_cost'])
        total_vat = float(data['total_vat'])
        total_sum = float(data['total_sum'])
        price = float(data['price'])

        items.append({
            'nom_id': data['nom_id'],
            'title': data['title'] or 'Без названия',
            'izm': data['izm'] or 'шт',
            'quantity': quantity,
            'price': price,  # ← реальная цена из документа
            'total_without_vat': total_cost,
            'vat_rate': float(data['vat_rate']),
            'total_vat': total_vat,
            'total': total_sum,
            'doc_id': data['doc_id'],
        })

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum

    # Форматирование дат
    date_start_display = date_start or ''
    date_end_display = date_end or ''

    context = {
        'object': obj,
        'items': items,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
        'grand_total': grand_total,
    }
    return render(request, 'inventory/reports/object_details.html', context)


def objects_report_excel(request):
    """Экспорт отчета по объектам в Excel"""
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    docs = Doc.objects.filter(oper=3).select_related('obct', 'obct__idpodraz')

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            docs = docs.filter(datadoc__gte=start_date)
        elif start_date:
            docs = docs.filter(update_date__date__gte=start_date)

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            docs = docs.filter(datadoc__lte=end_date)
        elif end_date:
            docs = docs.filter(update_date__date__lte=end_date)

    # Группировка по объектам
    objects_dict = {}
    for doc in docs:
        if not doc.obct:
            continue
        obct_id = doc.obct.id
        if obct_id not in objects_dict:
            objects_dict[obct_id] = {
                'title': doc.obct.title or 'Без названия',
                'podraz': doc.obct.idpodraz.title if doc.obct.idpodraz else '—',
                'total_without_vat': 0,
                'total_vat': 0,
                'total': 0
            }

        for detail in doc.details.all():
            objects_dict[obct_id]['total_without_vat'] += abs(detail.cost)
            objects_dict[obct_id]['total_vat'] += abs(detail.vat_amount)
            objects_dict[obct_id]['total'] += abs(detail.total_with_vat)

    objects = list(objects_dict.values())
    objects.sort(key=lambda x: x['title'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по объектам"

    headers = ['№ п/п', 'Объект', 'Подразделение', 'Стоимость без НДС', 'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, obj in enumerate(objects, 1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=obj['title'])
        ws.cell(row=row_num, column=3, value=obj['podraz'])
        ws.cell(row=row_num, column=4, value=obj['total_without_vat'])
        ws.cell(row=row_num, column=5, value=obj['total_vat'])
        ws.cell(row=row_num, column=6, value=obj['total'])
        row_num += 1
        grand_total_without_vat += obj['total_without_vat']
        grand_total_vat += obj['total_vat']
        grand_total += obj['total']

    ws.cell(row=row_num, column=3, value='ИТОГО:')
    ws.cell(row=row_num, column=4, value=grand_total_without_vat)
    ws.cell(row=row_num, column=5, value=grand_total_vat)
    ws.cell(row=row_num, column=6, value=grand_total)

    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=object_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def object_details_excel(request, object_id):
    """Экспорт детализации по объекту в Excel (список товаров)"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    obj = get_object_or_404(Obct, id=object_id)

    query = """
        SELECT 
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 3 AND doc.obct_id = %s
    """

    params = [object_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Детализация {obj.title[:30]}"

    headers = ['№ п/п', 'Наименование', 'Ед.изм.', 'Кол-во', 'Цена', 'Стоимость без НДС', 'НДС %', 'Сумма НДС',
               'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        title = row[0] or 'Без названия'
        izm = row[1] or 'шт'
        price = float(row[2])
        vat_rate = float(row[3])
        quantity = abs(float(row[4]))
        total_cost = abs(float(row[5]))
        total_vat = abs(float(row[6]))
        total_sum = abs(float(row[7]))

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=title)
        ws.cell(row=row_num, column=3, value=izm)
        ws.cell(row=row_num, column=4, value=quantity)
        ws.cell(row=row_num, column=5, value=price)
        ws.cell(row=row_num, column=6, value=total_cost)
        ws.cell(row=row_num, column=7, value=vat_rate)
        ws.cell(row=row_num, column=8, value=total_vat)
        ws.cell(row=row_num, column=9, value=total_sum)

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum
        row_num += 1

    ws.cell(row=row_num, column=6, value='ИТОГО:')
    ws.cell(row=row_num, column=7, value=grand_total_without_vat)
    ws.cell(row=row_num, column=8, value=grand_total_vat)
    ws.cell(row=row_num, column=9, value=grand_total)

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=object_details_{object_id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def departments_report(request):
    """Отчет по подразделениям (перемещения)"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_departments_report_table(request, date_start, date_end, date_type)

    return render(request, 'inventory/reports/department_report.html', {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    })


def render_departments_report_table(request, date_start, date_end, date_type='created'):
    """Рендеринг таблицы отчета по подразделениям"""
    from datetime import datetime, timedelta
    from django.db import connection

    # Исправлено: o.idpodraz вместо o.idpodraz_id
    query = """
           SELECT 
               p.id as podraz_id,
               p.title as podraz_title,
               SUM(ABS(d.cost)) as total_without_vat,
               SUM(ABS(d.vat_amount)) as total_vat,
               SUM(ABS(d.total_with_vat)) as total_sum
           FROM detail d
           INNER JOIN doc ON d.id_doc = doc.id
           INNER JOIN obct o ON doc.obct_id = o.id
           INNER JOIN podraz p ON o.idpodraz = p.id
           WHERE doc.oper = 3
       """

    params = []

    # Фильтрация по дате
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
           GROUP BY p.id, p.title
           ORDER BY p.title
       """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    departments = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        departments.append({
            'id': row[0],
            'title': row[1] or 'Без названия',
            'total_without_vat': float(row[2]),
            'total_vat': float(row[3]),
            'total': float(row[4]),
        })
        grand_total_without_vat += float(row[2])
        grand_total_vat += float(row[3])
        grand_total += float(row[4])

    # Форматирование дат
    date_start_display = ''
    date_end_display = ''
    if date_start:
        try:
            date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_start_display = date_start
    if date_end:
        try:
            date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_end_display = date_end

    html = render_to_string('inventory/reports/department_report_table.html', {
        'departments': departments,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'date_type': date_type,
        'grand_total': grand_total,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
    })
    return HttpResponse(html)


def department_details(request, department_id):
    """Детализация по подразделению"""
    from datetime import datetime, timedelta
    from django.db import connection

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    department = get_object_or_404(Podraz, id=department_id)

    # Исправлено: o.idpodraz вместо o.idpodraz_id
    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            o.title as object_title,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum,
            MIN(d.id_doc) as doc_id
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN obct o ON doc.obct_id = o.id
        INNER JOIN podraz p ON o.idpodraz = p.id
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        WHERE doc.oper = 3 AND p.id = %s
    """

    params = [department_id]

    # Фильтрация по дате
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, o.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    items = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        data = dict(zip(columns, row))
        quantity = float(data['quantity'])
        total_cost = float(data['total_cost'])
        total_vat = float(data['total_vat'])
        total_sum = float(data['total_sum'])
        price = float(data['price'])

        items.append({
            'nom_id': data['nom_id'],
            'title': data['title'] or 'Без названия',
            'izm': data['izm'] or 'шт',
            'object_title': data['object_title'] or '—',
            'quantity': quantity,
            'price': price,
            'total_without_vat': total_cost,
            'vat_rate': float(data['vat_rate']),
            'total_vat': total_vat,
            'total': total_sum,
            'doc_id': data['doc_id'],
        })

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum

    # Форматирование дат
    date_start_display = date_start or ''
    date_end_display = date_end or ''

    context = {
        'department': department,
        'items': items,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
        'grand_total': grand_total,
        'date_start': date_start_display,
        'date_end': date_end_display,
    }
    return render(request, 'inventory/reports/department_details.html', context)


def departments_report_excel(request):
    """Экспорт отчета по подразделениям в Excel"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    # Исправлено: o.idpodraz (без _id)
    query = """
        SELECT 
            p.title as podraz_title,
            SUM(ABS(d.cost)) as total_without_vat,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN obct o ON doc.obct_id = o.id
        INNER JOIN podraz p ON o.idpodraz = p.id
        WHERE doc.oper = 3
    """

    params = []

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY p.id, p.title
        ORDER BY p.title
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по подразделениям"

    headers = ['№ п/п', 'Подразделение', 'Стоимость без НДС', 'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=row[0] or 'Без названия')
        ws.cell(row=row_num, column=3, value=float(row[1]))
        ws.cell(row=row_num, column=4, value=float(row[2]))
        ws.cell(row=row_num, column=5, value=float(row[3]))

        grand_total_without_vat += float(row[1])
        grand_total_vat += float(row[2])
        grand_total += float(row[3])
        row_num += 1

    ws.cell(row=row_num, column=3, value='ИТОГО:')
    ws.cell(row=row_num, column=4, value=grand_total_without_vat)
    ws.cell(row=row_num, column=5, value=grand_total_vat)
    ws.cell(row=row_num, column=6, value=grand_total)

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=department_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def department_details_excel(request, department_id):
    """Экспорт детализации по подразделению в Excel"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    department = get_object_or_404(Podraz, id=department_id)

    # Исправлено: o.idpodraz (без _id)
    query = """
        SELECT 
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            o.title as object_title,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN obct o ON doc.obct_id = o.id
        INNER JOIN podraz p ON o.idpodraz = p.id
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        WHERE doc.oper = 3 AND p.id = %s
    """

    params = [department_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, o.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Детализация {department.title[:30]}"

    headers = ['№ п/п', 'Наименование', 'Ед.изм.', 'Объект', 'Кол-во', 'Цена', 'Стоимость без НДС', 'НДС %',
               'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        title = row[0] or 'Без названия'
        izm = row[1] or 'шт'
        object_title = row[2] or '—'
        price = float(row[3])
        vat_rate = float(row[4])
        quantity = abs(float(row[5]))
        total_cost = abs(float(row[6]))
        total_vat = abs(float(row[7]))
        total_sum = abs(float(row[8]))

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=title)
        ws.cell(row=row_num, column=3, value=izm)
        ws.cell(row=row_num, column=4, value=object_title)
        ws.cell(row=row_num, column=5, value=quantity)
        ws.cell(row=row_num, column=6, value=price)
        ws.cell(row=row_num, column=7, value=total_cost)
        ws.cell(row=row_num, column=8, value=vat_rate)
        ws.cell(row=row_num, column=9, value=total_vat)
        ws.cell(row=row_num, column=10, value=total_sum)

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum
        row_num += 1

    ws.cell(row=row_num, column=7, value='ИТОГО:')
    ws.cell(row=row_num, column=8, value=grand_total_without_vat)
    ws.cell(row=row_num, column=9, value=grand_total_vat)
    ws.cell(row=row_num, column=10, value=grand_total)

    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=department_details_{department_id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def fio_report(request):
    """Отчет по подотчетным лицам (перемещения)"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_fio_report_table(request, date_start, date_end, date_type)

    return render(request, 'inventory/reports/fio_report.html', {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    })


def render_fio_report_table(request, date_start, date_end, date_type='created'):
    """Рендеринг таблицы отчета по подотчетным лицам"""
    from datetime import datetime, timedelta
    from django.db import connection

    query = """
        SELECT 
            f.id as fio_id,
            f.title as fio_title,
            SUM(ABS(d.cost)) as total_without_vat,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN fio f ON doc.fio_id = f.id
        WHERE doc.oper = 3
    """

    params = []

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY f.id, f.title
        ORDER BY f.title
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    fios = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        fios.append({
            'id': row[0],
            'title': row[1] or 'Без названия',
            'total_without_vat': float(row[2]),
            'total_vat': float(row[3]),
            'total': float(row[4]),
        })
        grand_total_without_vat += float(row[2])
        grand_total_vat += float(row[3])
        grand_total += float(row[4])

    date_start_display = ''
    date_end_display = ''
    if date_start:
        try:
            date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_start_display = date_start
    if date_end:
        try:
            date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_end_display = date_end

    html = render_to_string('inventory/reports/fio_report_table.html', {
        'fios': fios,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'date_type': date_type,
        'grand_total': grand_total,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
    })
    return HttpResponse(html)


def fio_details(request, fio_id):
    """Детализация по подотчетному лицу"""
    from datetime import datetime, timedelta
    from django.db import connection

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    fio = get_object_or_404(Fio, id=fio_id)

    date_start_display = date_start or ''
    date_end_display = date_end or ''

    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            o.title as object_title,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum,
            MIN(d.id_doc) as doc_id
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN obct o ON doc.obct_id = o.id
        INNER JOIN fio f ON doc.fio_id = f.id
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        WHERE doc.oper = 3 AND f.id = %s
    """

    params = [fio_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, o.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    items = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        data = dict(zip(columns, row))
        quantity = float(data['quantity'])
        total_cost = float(data['total_cost'])
        total_vat = float(data['total_vat'])
        total_sum = float(data['total_sum'])
        price = float(data['price'])

        items.append({
            'nom_id': data['nom_id'],
            'title': data['title'] or 'Без названия',
            'izm': data['izm'] or 'шт',
            'object_title': data['object_title'] or '—',
            'quantity': quantity,
            'price': price,
            'total_without_vat': total_cost,
            'vat_rate': float(data['vat_rate']),
            'total_vat': total_vat,
            'total': total_sum,
            'doc_id': data['doc_id'],
        })

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum

    context = {
        'fio': fio,
        'items': items,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
        'grand_total': grand_total,
    }
    return render(request, 'inventory/reports/fio_details.html', context)


def fio_report_excel(request):
    """Экспорт отчета по подотчетным лицам в Excel"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    query = """
        SELECT 
            f.title as fio_title,
            SUM(ABS(d.cost)) as total_without_vat,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN fio f ON doc.fio_id = f.id
        WHERE doc.oper = 3
    """

    params = []

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY f.id, f.title
        ORDER BY f.title
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по подотчетным лицам"

    headers = ['№ п/п', 'Подотчетное лицо', 'Стоимость без НДС', 'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=row[0] or 'Без названия')
        ws.cell(row=row_num, column=3, value=float(row[1]))
        ws.cell(row=row_num, column=4, value=float(row[2]))
        ws.cell(row=row_num, column=5, value=float(row[3]))

        grand_total_without_vat += float(row[1])
        grand_total_vat += float(row[2])
        grand_total += float(row[3])
        row_num += 1

    ws.cell(row=row_num, column=3, value='ИТОГО:')
    ws.cell(row=row_num, column=4, value=grand_total_without_vat)
    ws.cell(row=row_num, column=5, value=grand_total_vat)
    ws.cell(row=row_num, column=6, value=grand_total)

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=fio_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def fio_details_excel(request, fio_id):
    """Экспорт детализации по подотчетному лицу в Excel"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    fio = get_object_or_404(Fio, id=fio_id)

    query = """
        SELECT 
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            o.title as object_title,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN doc ON d.id_doc = doc.id
        INNER JOIN obct o ON doc.obct_id = o.id
        INNER JOIN fio f ON doc.fio_id = f.id
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        WHERE doc.oper = 3 AND f.id = %s
    """

    params = [fio_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, o.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Детализация {fio.title[:30]}"

    headers = ['№ п/п', 'Наименование', 'Ед.изм.', 'Объект', 'Кол-во', 'Цена', 'Стоимость без НДС', 'НДС %',
               'Сумма НДС', 'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        title = row[0] or 'Без названия'
        izm = row[1] or 'шт'
        object_title = row[2] or '—'
        price = float(row[3])
        vat_rate = float(row[4])
        quantity = abs(float(row[5]))
        total_cost = abs(float(row[6]))
        total_vat = abs(float(row[7]))
        total_sum = abs(float(row[8]))

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=title)
        ws.cell(row=row_num, column=3, value=izm)
        ws.cell(row=row_num, column=4, value=object_title)
        ws.cell(row=row_num, column=5, value=quantity)
        ws.cell(row=row_num, column=6, value=price)
        ws.cell(row=row_num, column=7, value=total_cost)
        ws.cell(row=row_num, column=8, value=vat_rate)
        ws.cell(row=row_num, column=9, value=total_vat)
        ws.cell(row=row_num, column=10, value=total_sum)

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum
        row_num += 1

    ws.cell(row=row_num, column=7, value='ИТОГО:')
    ws.cell(row=row_num, column=8, value=grand_total_without_vat)
    ws.cell(row=row_num, column=9, value=grand_total_vat)
    ws.cell(row=row_num, column=10, value=grand_total)

    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=fio_details_{fio_id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def categories_report(request):
    """Отчет по категориям ТМЦ (приход и расход)"""
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_categories_report_table(request, date_start, date_end, date_type)

    return render(request, 'inventory/reports/category_report.html', {
        'date_start': date_start,
        'date_end': date_end,
        'date_type': date_type,
    })


def render_categories_report_table(request, date_start, date_end, date_type='created'):
    """Рендеринг таблицы отчета по категориям с начальными остатками"""
    from datetime import datetime, timedelta
    from django.db import connection

    # Приход (oper=2)
    query_income = """
        SELECT 
            c.id as category_id,
            c.title as category_title,
            SUM(ABS(d.total_with_vat)) as total_income
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        INNER JOIN category c ON n.category_id = c.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 2
    """

    # Расход (oper=3)
    query_outcome = """
        SELECT 
            c.id as category_id,
            c.title as category_title,
            SUM(ABS(d.total_with_vat)) as total_outcome
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        INNER JOIN category c ON n.category_id = c.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 3
    """

    # Начальные остатки (oper=1)
    query_initial = """
        SELECT 
            c.id as category_id,
            c.title as category_title,
            SUM(ABS(d.total_with_vat)) as total_initial
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        INNER JOIN category c ON n.category_id = c.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 1
    """

    params = []

    # Фильтрация по дате для всех запросов
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query_income += " AND doc.datadoc >= %s"
            query_outcome += " AND doc.datadoc >= %s"
            query_initial += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query_income += " AND DATE(doc.update_date) >= %s"
            query_outcome += " AND DATE(doc.update_date) >= %s"
            query_initial += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query_income += " AND doc.datadoc <= %s"
            query_outcome += " AND doc.datadoc <= %s"
            query_initial += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query_income += " AND DATE(doc.update_date) <= %s"
            query_outcome += " AND DATE(doc.update_date) <= %s"
            query_initial += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query_income += " GROUP BY c.id, c.title"
    query_outcome += " GROUP BY c.id, c.title"
    query_initial += " GROUP BY c.id, c.title"

    # Выполняем запросы
    with connection.cursor() as cursor:
        cursor.execute(query_income, params)
        income_rows = {row[0]: {'title': row[1], 'total': float(row[2])} for row in cursor.fetchall()}

        cursor.execute(query_outcome, params)
        outcome_rows = {row[0]: {'title': row[1], 'total': float(row[2])} for row in cursor.fetchall()}

        cursor.execute(query_initial, params)
        initial_rows = {row[0]: {'title': row[1], 'total': float(row[2])} for row in cursor.fetchall()}

    # Объединяем результаты
    all_categories = set(income_rows.keys()) | set(outcome_rows.keys()) | set(initial_rows.keys())
    categories = []
    grand_total_income = 0
    grand_total_outcome = 0
    grand_total_initial = 0

    for cat_id in all_categories:
        cat_title = income_rows.get(cat_id, {}).get('title') or \
                    outcome_rows.get(cat_id, {}).get('title') or \
                    initial_rows.get(cat_id, {}).get('title') or 'Без категории'
        total_income = income_rows.get(cat_id, {}).get('total', 0)
        total_outcome = outcome_rows.get(cat_id, {}).get('total', 0)
        total_initial = initial_rows.get(cat_id, {}).get('total', 0)

        categories.append({
            'id': cat_id,
            'title': cat_title,
            'total_income': total_income,
            'total_outcome': total_outcome,
            'total_initial': total_initial,
        })
        grand_total_income += total_income
        grand_total_outcome += total_outcome
        grand_total_initial += total_initial

    categories.sort(key=lambda x: x['title'])

    # Форматирование дат
    date_start_display = ''
    date_end_display = ''
    if date_start:
        try:
            date_start_display = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_start_display = date_start
    if date_end:
        try:
            date_end_display = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_end_display = date_end

    html = render_to_string('inventory/reports/category_report_table.html', {
        'categories': categories,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'date_type': date_type,
        'grand_total_income': grand_total_income,
        'grand_total_outcome': grand_total_outcome,
        'grand_total_initial': grand_total_initial,
    })
    return HttpResponse(html)


def category_income_details(request, category_id):
    """Детализация по приходу категории (без группировки)"""
    from datetime import datetime, timedelta
    from django.db import connection

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    category = get_object_or_404(Category, id=category_id)

    date_start_display = date_start or ''
    date_end_display = date_end or ''

    # Убираем GROUP BY, показываем каждую строку отдельно
    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            d.vat_rate,
            d.kolvo as quantity,
            ABS(d.cost) as total_cost,
            ABS(d.vat_amount) as total_vat,
            ABS(d.total_with_vat) as total_sum,
            d.id_doc as doc_id,
            doc.nomer as doc_nomer,
            doc.datadoc as doc_date
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 2 AND n.category_id = %s
    """

    params = [category_id]

    # Фильтрация по дате
    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    # Убираем GROUP BY, сортируем по дате документа
    query += " ORDER BY doc.datadoc, n.title"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    items = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        data = dict(zip(columns, row))
        quantity = float(data['quantity'])
        total_cost = float(data['total_cost'])
        total_vat = float(data['total_vat'])
        total_sum = float(data['total_sum'])
        price = float(data['price'])

        items.append({
            'nom_id': data['nom_id'],
            'title': data['title'] or 'Без названия',
            'izm': data['izm'] or 'шт',
            'quantity': quantity,
            'price': price,
            'total_without_vat': total_cost,
            'vat_rate': float(data['vat_rate']),
            'total_vat': total_vat,
            'total': total_sum,
            'doc_id': data['doc_id'],
            'doc_nomer': data['doc_nomer'],
            'doc_date': data['doc_date'],
        })

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum

    context = {
        'category': category,
        'items': items,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
        'grand_total': grand_total,
    }
    return render(request, 'inventory/reports/category_income_details.html', context)


def category_outcome_details(request, category_id):
    """Детализация по расходу категории (без группировки)"""
    from datetime import datetime, timedelta
    from django.db import connection

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    category = get_object_or_404(Category, id=category_id)

    date_start_display = date_start or ''
    date_end_display = date_end or ''

    query = """
        SELECT 
            n.id as nom_id,
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            d.vat_rate,
            ABS(d.kolvo) as quantity,
            ABS(d.cost) as total_cost,
            ABS(d.vat_amount) as total_vat,
            ABS(d.total_with_vat) as total_sum,
            d.id_doc as doc_id,
            doc.nomer as doc_nomer,
            doc.datadoc as doc_date
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 3 AND n.category_id = %s
    """

    params = [category_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += " ORDER BY doc.datadoc, n.title"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    items = []
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for row in rows:
        data = dict(zip(columns, row))
        quantity = float(data['quantity'])
        total_cost = float(data['total_cost'])
        total_vat = float(data['total_vat'])
        total_sum = float(data['total_sum'])
        price = float(data['price'])

        items.append({
            'nom_id': data['nom_id'],
            'title': data['title'] or 'Без названия',
            'izm': data['izm'] or 'шт',
            'quantity': quantity,
            'price': price,
            'total_without_vat': total_cost,
            'vat_rate': float(data['vat_rate']),
            'total_vat': total_vat,
            'total': total_sum,
            'doc_id': data['doc_id'],
            'doc_nomer': data['doc_nomer'],
            'doc_date': data['doc_date'],
        })

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum

    context = {
        'category': category,
        'items': items,
        'date_start': date_start_display,
        'date_end': date_end_display,
        'grand_total_without_vat': grand_total_without_vat,
        'grand_total_vat': grand_total_vat,
        'grand_total': grand_total,
    }
    return render(request, 'inventory/reports/category_outcome_details.html', context)


def categories_report_excel(request):
    """Экспорт отчета по категориям в Excel (с начальными остатками)"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    # Приход
    query_income = """
        SELECT 
            c.title as category_title,
            SUM(ABS(d.total_with_vat)) as total_income
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        INNER JOIN category c ON n.category_id = c.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 2
    """

    # Расход
    query_outcome = """
        SELECT 
            c.title as category_title,
            SUM(ABS(d.total_with_vat)) as total_outcome
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        INNER JOIN category c ON n.category_id = c.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 3
    """

    # Начальные остатки
    query_initial = """
        SELECT 
            c.title as category_title,
            SUM(ABS(d.total_with_vat)) as total_initial
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        INNER JOIN category c ON n.category_id = c.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 1
    """

    params = []

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%Y-%m-%d').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query_income += " AND doc.datadoc >= %s"
            query_outcome += " AND doc.datadoc >= %s"
            query_initial += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query_income += " AND DATE(doc.update_date) >= %s"
            query_outcome += " AND DATE(doc.update_date) >= %s"
            query_initial += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%Y-%m-%d').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query_income += " AND doc.datadoc <= %s"
            query_outcome += " AND doc.datadoc <= %s"
            query_initial += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query_income += " AND DATE(doc.update_date) <= %s"
            query_outcome += " AND DATE(doc.update_date) <= %s"
            query_initial += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query_income += " GROUP BY c.id, c.title ORDER BY c.title"
    query_outcome += " GROUP BY c.id, c.title ORDER BY c.title"
    query_initial += " GROUP BY c.id, c.title ORDER BY c.title"

    with connection.cursor() as cursor:
        cursor.execute(query_income, params)
        income_rows = {row[0]: float(row[1]) for row in cursor.fetchall()}

        cursor.execute(query_outcome, params)
        outcome_rows = {row[0]: float(row[1]) for row in cursor.fetchall()}

        cursor.execute(query_initial, params)
        initial_rows = {row[0]: float(row[1]) for row in cursor.fetchall()}

    all_categories = set(income_rows.keys()) | set(outcome_rows.keys()) | set(initial_rows.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по категориям"

    headers = ['№ п/п', 'Категория', 'Нач.остатки', 'Получено', 'Выдано']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_initial = 0
    grand_total_income = 0
    grand_total_outcome = 0

    for i, cat_title in enumerate(sorted(all_categories), 1):
        total_initial = initial_rows.get(cat_title, 0)
        total_income = income_rows.get(cat_title, 0)
        total_outcome = outcome_rows.get(cat_title, 0)

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=cat_title or 'Без категории')
        ws.cell(row=row_num, column=3, value=total_initial)
        ws.cell(row=row_num, column=4, value=total_income)
        ws.cell(row=row_num, column=5, value=total_outcome)

        grand_total_initial += total_initial
        grand_total_income += total_income
        grand_total_outcome += total_outcome
        row_num += 1

    ws.cell(row=row_num, column=3, value='ИТОГО:')
    ws.cell(row=row_num, column=4, value=grand_total_initial)
    ws.cell(row=row_num, column=5, value=grand_total_income)
    ws.cell(row=row_num, column=6, value=grand_total_outcome)

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=category_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def category_income_details_excel(request, category_id):
    """Экспорт детализации по приходу категории в Excel"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    category = get_object_or_404(Category, id=category_id)

    query = """
        SELECT 
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            d.vat_rate,
            SUM(d.kolvo) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 2 AND n.category_id = %s
    """

    params = [category_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Приход {category.title[:25]}"

    headers = ['№ п/п', 'Наименование', 'Ед.изм.', 'Кол-во', 'Цена', 'Стоимость без НДС', 'НДС %', 'Сумма НДС',
               'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        title = row[0] or 'Без названия'
        izm = row[1] or 'шт'
        price = float(row[2])
        vat_rate = float(row[3])
        quantity = float(row[4])
        total_cost = abs(float(row[5]))
        total_vat = abs(float(row[6]))
        total_sum = abs(float(row[7]))

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=title)
        ws.cell(row=row_num, column=3, value=izm)
        ws.cell(row=row_num, column=4, value=quantity)
        ws.cell(row=row_num, column=5, value=price)
        ws.cell(row=row_num, column=6, value=total_cost)
        ws.cell(row=row_num, column=7, value=vat_rate)
        ws.cell(row=row_num, column=8, value=total_vat)
        ws.cell(row=row_num, column=9, value=total_sum)

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum
        row_num += 1

    ws.cell(row=row_num, column=6, value='ИТОГО:')
    ws.cell(row=row_num, column=7, value=grand_total_without_vat)
    ws.cell(row=row_num, column=8, value=grand_total_vat)
    ws.cell(row=row_num, column=9, value=grand_total)

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=category_income_{category_id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response


def category_outcome_details_excel(request, category_id):
    """Экспорт детализации по расходу категории в Excel"""
    from datetime import datetime
    from django.db import connection
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    date_type = request.GET.get('date_type', 'created')

    category = get_object_or_404(Category, id=category_id)

    query = """
        SELECT 
            n.title as title,
            COALESCE(i.title, 'шт') as izm,
            d.price,
            d.vat_rate,
            SUM(ABS(d.kolvo)) as quantity,
            SUM(ABS(d.cost)) as total_cost,
            SUM(ABS(d.vat_amount)) as total_vat,
            SUM(ABS(d.total_with_vat)) as total_sum
        FROM detail d
        INNER JOIN nom n ON d.id_nom = n.id
        LEFT JOIN izm i ON n.izm_id = i.id
        INNER JOIN doc ON d.id_doc = doc.id
        WHERE doc.oper = 3 AND n.category_id = %s
    """

    params = [category_id]

    if date_start:
        try:
            start_date = datetime.strptime(date_start, '%d.%m.%Y').date()
        except:
            start_date = None

        if date_type == 'doc' and start_date:
            query += " AND doc.datadoc >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))
        elif start_date:
            query += " AND DATE(doc.update_date) >= %s"
            params.append(start_date.strftime('%Y-%m-%d'))

    if date_end:
        try:
            end_date = datetime.strptime(date_end, '%d.%m.%Y').date()
        except:
            end_date = None

        if date_type == 'doc' and end_date:
            query += " AND doc.datadoc <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))
        elif end_date:
            query += " AND DATE(doc.update_date) <= %s"
            params.append(end_date.strftime('%Y-%m-%d'))

    query += """
        GROUP BY n.id, n.title, i.title, d.price, d.vat_rate
        HAVING SUM(d.kolvo) != 0
        ORDER BY n.title, d.price
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Расход {category.title[:25]}"

    headers = ['№ п/п', 'Наименование', 'Ед.изм.', 'Кол-во', 'Цена', 'Стоимость без НДС', 'НДС %', 'Сумма НДС',
               'Стоимость с НДС']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    grand_total_without_vat = 0
    grand_total_vat = 0
    grand_total = 0

    for i, row in enumerate(rows, 1):
        title = row[0] or 'Без названия'
        izm = row[1] or 'шт'
        price = float(row[2])
        vat_rate = float(row[3])
        quantity = float(row[4])
        total_cost = abs(float(row[5]))
        total_vat = abs(float(row[6]))
        total_sum = abs(float(row[7]))

        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=title)
        ws.cell(row=row_num, column=3, value=izm)
        ws.cell(row=row_num, column=4, value=quantity)
        ws.cell(row=row_num, column=5, value=price)
        ws.cell(row=row_num, column=6, value=total_cost)
        ws.cell(row=row_num, column=7, value=vat_rate)
        ws.cell(row=row_num, column=8, value=total_vat)
        ws.cell(row=row_num, column=9, value=total_sum)

        grand_total_without_vat += total_cost
        grand_total_vat += total_vat
        grand_total += total_sum
        row_num += 1

    ws.cell(row=row_num, column=6, value='ИТОГО:')
    ws.cell(row=row_num, column=7, value=grand_total_without_vat)
    ws.cell(row=row_num, column=8, value=grand_total_vat)
    ws.cell(row=row_num, column=9, value=grand_total)

    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=category_outcome_{category_id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return response




from django.db.models import Sum, Max, Value, DecimalField
from django.db.models.functions import Coalesce, Abs
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string



def top10_report(request):
    """Отчет ТОП-5 за период (основная страница)"""
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    date_field = request.GET.get('date_field', 'datadoc')

    # Формируем фильтр по датам (если даты заданы)
    date_filter = {}
    if from_date and to_date:
        if date_field == 'datadoc':
            date_filter = {
                'id_doc__datadoc__gte': from_date,
                'id_doc__datadoc__lte': to_date,
            }
        else:  # update_date
            date_filter = {
                'id_doc__update_date__date__gte': from_date,
                'id_doc__update_date__date__lte': to_date,
            }

    # Базовый queryset (исключаем oper=1 - начальные остатки)
    details = Detail.objects.filter(**date_filter).exclude(id_doc__oper=1)

    # 1. ТОП-5 поставщиков по сумме (только oper=2 - поступления)
    top_suppliers = (Detail.objects
                     .filter(**date_filter, id_doc__oper=2, id_doc__postav__isnull=False)
                     .values('id_doc__postav__title')
                     .annotate(total_sum=Coalesce(Sum('total_with_vat'), Value(0, output_field=DecimalField())))
                     .filter(total_sum__gt=0)
                     .order_by('-total_sum')[:5])

    # 2. ТОП-5 объектов по расходу (модуль)
    top_objects = (Detail.objects
                   .filter(**date_filter, id_doc__oper=3, id_doc__obct__isnull=False)
                   .values('id_doc__obct__title')
                   .annotate(total_sum=Coalesce(Abs(Sum('total_with_vat')), Value(0, output_field=DecimalField())))
                   .filter(total_sum__gt=0)
                   .order_by('-total_sum')[:5])

    # 3. ТОП-5 подразделений по расходу (модуль)
    top_departments = (Detail.objects
                       .filter(**date_filter, id_doc__oper=3, id_doc__obct__idpodraz__isnull=False)
                       .values('id_doc__obct__idpodraz__title')
                       .annotate(total_sum=Coalesce(Abs(Sum('total_with_vat')), Value(0, output_field=DecimalField())))
                       .filter(total_sum__gt=0)
                       .order_by('-total_sum')[:5])

    # 4. ТОП-5 подотчетных лиц по расходу (модуль)
    top_fio = (Detail.objects
               .filter(**date_filter, id_doc__oper=3, id_doc__fio__isnull=False)
               .values('id_doc__fio__title')
               .annotate(total_sum=Coalesce(Abs(Sum('total_with_vat')), Value(0, output_field=DecimalField())))
               .filter(total_sum__gt=0)
               .order_by('-total_sum')[:5])

    # 5. ТОП-5 самых дорогих ПРИОБРЕТЕННЫХ товаров (oper=2 - поступление)
    top_expensive_purchased = (Detail.objects
                               .filter(**date_filter, id_doc__oper=2)
                               .values('id_nom__title', 'id_nom__izm__title')
                               .annotate(max_price=Coalesce(Max('price'), Value(0, output_field=DecimalField())))
                               .filter(max_price__gt=0)
                               .order_by('-max_price')[:5])

    # 6. ТОП-5 самых дорогих ВЫДАННЫХ товаров (oper=3 - перемещение)
    top_expensive_issued = (Detail.objects
                            .filter(**date_filter, id_doc__oper=3)
                            .values('id_nom__title', 'id_nom__izm__title')
                            .annotate(max_price=Coalesce(Max('price'), Value(0, output_field=DecimalField())))
                            .filter(max_price__gt=0)
                            .order_by('-max_price')[:5])

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'date_field': date_field,
        'top_suppliers': top_suppliers,
        'top_objects': top_objects,
        'top_departments': top_departments,
        'top_fio': top_fio,
        'top_expensive_purchased': top_expensive_purchased,
        'top_expensive_issued': top_expensive_issued,
    }

    return render(request, 'inventory/reports/top_10_report.html', context)


def top10_report_data(request):
    """AJAX-вьюха для получения данных ТОП-5"""
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    date_field = request.GET.get('date_field', 'datadoc')

    # Формируем фильтр по датам (если даты заданы)
    date_filter = {}
    if from_date and to_date:
        if date_field == 'datadoc':
            date_filter = {
                'id_doc__datadoc__gte': from_date,
                'id_doc__datadoc__lte': to_date,
            }
        else:  # update_date
            date_filter = {
                'id_doc__update_date__date__gte': from_date,
                'id_doc__update_date__date__lte': to_date,
            }

    # Базовый queryset (исключаем oper=1 - начальные остатки)
    details = Detail.objects.filter(**date_filter).exclude(id_doc__oper=1)

    # 1. ТОП-5 поставщиков по сумме (только oper=2 - поступления)
    top_suppliers = (Detail.objects
                     .filter(**date_filter, id_doc__oper=2, id_doc__postav__isnull=False)
                     .values('id_doc__postav__title')
                     .annotate(total_sum=Coalesce(Sum('total_with_vat'), Value(0, output_field=DecimalField())))
                     .filter(total_sum__gt=0)
                     .order_by('-total_sum')[:5])

    # 2. ТОП-5 объектов по расходу (модуль)
    top_objects = (Detail.objects
                   .filter(**date_filter, id_doc__oper=3, id_doc__obct__isnull=False)
                   .values('id_doc__obct__title')
                   .annotate(total_sum=Coalesce(Abs(Sum('total_with_vat')), Value(0, output_field=DecimalField())))
                   .filter(total_sum__gt=0)
                   .order_by('-total_sum')[:5])

    # 3. ТОП-5 подразделений по расходу (модуль)
    top_departments = (Detail.objects
                       .filter(**date_filter, id_doc__oper=3, id_doc__obct__idpodraz__isnull=False)
                       .values('id_doc__obct__idpodraz__title')
                       .annotate(total_sum=Coalesce(Abs(Sum('total_with_vat')), Value(0, output_field=DecimalField())))
                       .filter(total_sum__gt=0)
                       .order_by('-total_sum')[:5])

    # 4. ТОП-5 подотчетных лиц по расходу (модуль)
    top_fio = (Detail.objects
               .filter(**date_filter, id_doc__oper=3, id_doc__fio__isnull=False)
               .values('id_doc__fio__title')
               .annotate(total_sum=Coalesce(Abs(Sum('total_with_vat')), Value(0, output_field=DecimalField())))
               .filter(total_sum__gt=0)
               .order_by('-total_sum')[:5])

    # 5. ТОП-5 самых дорогих ПРИОБРЕТЕННЫХ товаров (oper=2 - поступление)
    top_expensive_purchased = (Detail.objects
                               .filter(**date_filter, id_doc__oper=2)
                               .values('id_nom__title', 'id_nom__izm__title')
                               .annotate(max_price=Coalesce(Max('price'), Value(0, output_field=DecimalField())))
                               .filter(max_price__gt=0)
                               .order_by('-max_price')[:5])

    # 6. ТОП-5 самых дорогих ВЫДАННЫХ товаров (oper=3 - перемещение)
    top_expensive_issued = (Detail.objects
                            .filter(**date_filter, id_doc__oper=3)
                            .values('id_nom__title', 'id_nom__izm__title')
                            .annotate(max_price=Coalesce(Max('price'), Value(0, output_field=DecimalField())))
                            .filter(max_price__gt=0)
                            .order_by('-max_price')[:5])

    context = {
        'from_date': from_date,
        'to_date': to_date,
        'date_field': date_field,
        'top_suppliers': top_suppliers,
        'top_objects': top_objects,
        'top_departments': top_departments,
        'top_fio': top_fio,
        'top_expensive_purchased': top_expensive_purchased,
        'top_expensive_issued': top_expensive_issued,
    }

    html = render_to_string('inventory/reports/top_5_content.html', context)
    return HttpResponse(html)